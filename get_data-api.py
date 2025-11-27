# get_data-api.py

import argparse
import csv
import os
import requests
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from loguru import logger
import certifi

# Конфигурация API
from conf import BASE_URL, HEADERS

BASE_URL = BASE_URL.rstrip('/')

# Настройка логгера
logger.remove()
logger.add(
    "logs/get_data_api.log",
    level="DEBUG",
    format="{time:YYYY-MM-DD HH:mm:ss.SSS} | {level} | {module}:{function}:{line} - {message}"
)

# Парсинг аргументов командной строки
parser = argparse.ArgumentParser()
parser.add_argument('--update-ids', type=str, help='Файл с ID товаров для обновления')
parser.add_argument('--single-id', type=int, help='ID одного товара для обновления')
args = parser.parse_args()


def create_report():
    logger.info("Создаём новый отчёт товаров через Ozon API")
    payload = {
        "language": "DEFAULT", 
        "offer_id": [], 
        "search": "",
        "sku": [], 
        "visibility": "ALL"
    }
    
    try:
        resp = requests.post(
            f"{BASE_URL}/v1/report/products/create",
            headers=HEADERS, 
            json=payload, 
            verify=certifi.where()
        )
        logger.debug(f"POST /v1/report/products/create → {resp.status_code}")
        resp.raise_for_status()
        
        report_code = resp.json()["result"]["code"]
        logger.info(f"Отчёт создан успешно, код: {report_code}")
        return report_code
        
    except Exception as e:
        logger.error(f"Ошибка при создании отчёта: {e}")
        raise


def check_report_status(code):
    logger.info(f"Ожидаем готовность отчёта {code}")
    payload = {"code": code}
    
    for attempt in range(1, 21):
        try:
            resp = requests.post(
                f"{BASE_URL}/v1/report/info",
                headers=HEADERS, 
                json=payload, 
                verify=certifi.where()
            )
            logger.debug(f"POST /v1/report/info → {resp.status_code} (попытка {attempt}/20)")
            resp.raise_for_status()
            
            res = resp.json()["result"]
            status = res['status']
            logger.debug(f"Статус отчёта: {status}")
            
            if status == "success":
                logger.info(f"Отчёт готов после {attempt} попыток")
                return res["file"]
            elif status in ("error", "expired"):
                logger.error(f"Отчёт завершился с ошибкой: {status}")
                return None
                
            logger.debug(f"Отчёт ещё не готов, ждём 10 секунд...")
            time.sleep(10)
            
        except Exception as e:
            logger.error(f"Ошибка при проверке статуса отчёта (попытка {attempt}): {e}")
            time.sleep(10)
    
    logger.error("Превышено время ожидания готовности отчёта")
    return None


def download_report(path):
    if not path:
        logger.error("Получен пустой путь к файлу отчёта")
        raise RuntimeError("Не удалось получить путь к файлу отчёта")
        
    # Формируем полный URL если получен относительный путь
    if not path.startswith("http"):
        path = f"https://cdn1.ozone.ru/s3/{path.lstrip('/')}"
        
    logger.info(f"Скачиваем отчёт по адресу: {path}")
    
    try:
        resp = requests.get(path, headers=HEADERS, verify=certifi.where())
        logger.debug(f"GET {path} → {resp.status_code}")
        resp.raise_for_status()
        
        # Декодируем содержимое с учётом BOM
        txt = resp.content.decode("utf-8-sig")
        
        # Парсим CSV с разделителем ";"
        products_data = list(csv.DictReader(txt.splitlines(), delimiter=";", quotechar='"'))
        logger.info(f"Успешно загружено {len(products_data)} товаров из отчёта")
        
        # Логируем доступные колонки для отладки
        if products_data:
            logger.debug(f"Доступные колонки в отчёте: {list(products_data[0].keys())}")
        
        return products_data
        
    except Exception as e:
        logger.error(f"Ошибка при скачивании отчёта: {e}")
        raise


def get_product_prices(product_ids):
    logger.info(f"Получаем цены для {len(product_ids)} товаров")
    prices = {}
    endpoint = "/v3/product/info/list"
    chunk_size = 100
    
    total_chunks = (len(product_ids) + chunk_size - 1) // chunk_size
    
    for i in range(0, len(product_ids), chunk_size):
        chunk = product_ids[i:i+chunk_size]
        chunk_num = i // chunk_size + 1
        
        logger.debug(f"Обрабатываем пакет {chunk_num}/{total_chunks} ({len(chunk)} товаров)")
        
        payload = {"product_id": [str(pid) for pid in chunk]}
        
        try:
            response = requests.post(
                f"{BASE_URL}{endpoint}",
                headers=HEADERS,
                json=payload,
                verify=certifi.where()
            )
            
            if response.status_code != 200:
                logger.error(f"Ошибка API для пакета {chunk_num}: {response.status_code} - {response.text}")
                continue
                
            data = response.json().get("items", [])
            
            for item in data:
                product_id = item.get("id")
                if product_id:
                    price_info = {
                        "base_price": item.get("price", "Н/Д"),
                        "old_price": item.get("old_price", "Н/Д"),
                        "marketing_price": item.get("marketing_price", "Н/Д"),
                        "min_price": item.get("min_price", "Н/Д"),
                        "currency": item.get("currency_code", "RUB")
                    }
                    prices[product_id] = price_info
                    
            logger.debug(f"Пакет {chunk_num} обработан, получено {len([x for x in data if x.get('id')])} цен")
            
        except Exception as e:
            logger.error(f"Ошибка при обработке пакета {chunk_num}: {e}")
            
        # Задержка между запросами для соблюдения rate limits
        time.sleep(0.5)
    
    logger.info(f"Получено цен для {len(prices)} товаров из {len(product_ids)} запрошенных")
    return prices


def extract_ids_from_report(products_data):
    """Извлекает ID товаров из данных отчёта для последующих API запросов"""
    product_ids = []
    
    for product in products_data:
        product_id_str = product.get("Ozon Product ID")
        if product_id_str and str(product_id_str).strip() and str(product_id_str).strip().isdigit():
            product_ids.append(int(product_id_str))
    
    logger.info(f"Извлечено {len(product_ids)} ID товаров из отчёта")
    return product_ids


def enrich_products_with_api_data(products_data):
    """Обогащает данные из отчёта информацией из API (цены)"""
    if not products_data:
        return products_data
    
    # Извлекаем ID для API запросов
    product_ids = extract_ids_from_report(products_data)
    
    if not product_ids:
        logger.warning("Не найдено ID товаров для запроса дополнительных данных")
        return products_data
    
    # Получаем цены через API
    prices = get_product_prices(product_ids)
    
    # Обогащаем данные ценами
    enriched_count = 0
    for product in products_data:
        product_id_str = product.get("Ozon Product ID")
        if product_id_str and str(product_id_str).strip() and str(product_id_str).strip().isdigit():
            product_id = int(product_id_str)
            if product_id in prices:
                product.update(prices[product_id])
                enriched_count += 1
    
    logger.info(f"Обогащено {enriched_count} товаров данными из API")
    return products_data


def load_opt_prices(opt_price_file="in/opt_all.xlsx"):
    logger.info(f"Загружаем цены из файла: {opt_price_file}")
    
    price_indexes = {
        'by_article': {},
        'by_code1c': {},
        'by_name': {},
        'by_nomenclature': {}
    }
    
    if not os.path.isfile(opt_price_file):
        logger.warning(f"Файл с ценами не найден: {opt_price_file}")
        return price_indexes
        
    try:
        wb_opt = load_workbook(opt_price_file, read_only=True)
        ws_opt = wb_opt.active
        
        if ws_opt is None:
            logger.error("Лист opt_price пуст")
            return price_indexes
           
        # Определение индексов колонок
        headers_opt = [cell.value for cell in ws_opt[1]]
        logger.debug(f"Заголовки в файле цен: {headers_opt}")
        
        # Поиск индексов нужных колонок
        column_indexes = {}
        search_columns = ['Артикул', 'Код 1С', 'Название товара', 'Номенклатура', 'Цена']
        
        for col_name in search_columns:
            if col_name in headers_opt:
                column_indexes[col_name] = headers_opt.index(col_name)
                logger.debug(f"Найдена колонка '{col_name}' с индексом {column_indexes[col_name]}")
            else:
                logger.debug(f"Колонка '{col_name}' не найдена в файле")
        
        if 'Цена' not in column_indexes:
            logger.error("Не найдена колонка 'Цена' в файле opt_price")
            return price_indexes
            
        price_idx = column_indexes['Цена']
        processed_rows = 0
        
        # Обрабатываем строки данных
        for row in ws_opt.iter_rows(min_row=2, values_only=True):
            if len(row) <= price_idx:
                continue
                
            price = row[price_idx]
            if price is None:
                continue
                
            processed_rows += 1
            
            # Индексируем по артикулу
            if 'Артикул' in column_indexes:
                art = row[column_indexes['Артикул']]
                if art:
                    art_clean = str(art).strip()
                    if art_clean:
                        price_indexes['by_article'][art_clean] = price
            
            # Индексируем по коду 1С
            if 'Код 1С' in column_indexes:
                code1c = row[column_indexes['Код 1С']]
                if code1c:
                    code1c_clean = str(code1c).strip()
                    if code1c_clean:
                        price_indexes['by_code1c'][code1c_clean] = price
            
            # Индексируем по названию товара
            if 'Название товара' in column_indexes:
                name = row[column_indexes['Название товара']]
                if name:
                    name_clean = str(name).strip()
                    if name_clean:
                        price_indexes['by_name'][name_clean] = price
            
            # Индексируем по номенклатуре
            if 'Номенклатура' in column_indexes:
                nom = row[column_indexes['Номенклатура']]
                if nom:
                    nom_clean = str(nom).strip()
                    if nom_clean:
                        price_indexes['by_nomenclature'][nom_clean] = price
        
        wb_opt.close()
        
        logger.info(f"Обработано {processed_rows} строк из файла цен")
        logger.info(f"Создано индексов: артикулы={len(price_indexes['by_article'])}, "
                   f"коды 1С={len(price_indexes['by_code1c'])}, "
                   f"названия={len(price_indexes['by_name'])}, "
                   f"номенклатура={len(price_indexes['by_nomenclature'])}")
               
    except Exception as e:
        logger.error(f"Ошибка при загрузке файла цен: {e}")
        
    return price_indexes


def find_price_for_product(item, price_indexes):
    article = str(item.get('Артикул', '')).strip()
    product_name = str(item.get('Название товара', '')).strip()
    
    # Этап 1: Поиск по артикулу (точное совпадение)
    if article:
        # Поиск в колонке "Артикул"
        if article in price_indexes['by_article']:
            logger.debug(f"Найдена цена по артикулу '{article}' в колонке Артикул")
            return price_indexes['by_article'][article]
        
        # Поиск в колонке "Код 1С"
        if article in price_indexes['by_code1c']:
            logger.debug(f"Найдена цена по артикулу '{article}' в колонке Код 1С")
            return price_indexes['by_code1c'][article]
    
    # Этап 2: Поиск по названию товара (точное совпадение)
    if product_name:
        # Поиск в колонке "Название товара"
        if product_name in price_indexes['by_name']:
            logger.debug(f"Найдена цена по названию '{product_name}' в колонке Название товара")
            return price_indexes['by_name'][product_name]
        
        # Поиск в колонке "Номенклатура"
        if product_name in price_indexes['by_nomenclature']:
            logger.debug(f"Найдена цена по названию '{product_name}' в колонке Номенклатура")
            return price_indexes['by_nomenclature'][product_name]
    
    return None


def enrich_products_with_prices(data, opt_price_file="in/opt_all.xlsx"):
    logger.info(f"Начинаем обогащение {len(data)} товаров ценами из {opt_price_file}")
    
    # Загружаем индексы цен
    price_indexes = load_opt_prices(opt_price_file)
    
    # Счетчики для статистики
    found_prices = 0
    not_found_prices = 0
    
    # Обрабатываем каждый товар
    for item in data:
        price = find_price_for_product(item, price_indexes)
        
        if price is not None:
            found_prices += 1
            item['Цена'] = price
        else:
            not_found_prices += 1
            item['Цена'] = None
            
            # Логируем товары без найденной цены для отладки
            article = item.get('Артикул', 'Н/Д')
            name = item.get('Название товара', 'Н/Д')
            logger.debug(f"Цена не найдена для товара: артикул='{article}', название='{name}'")
    
    logger.info(f"Результаты обогащения ценами: найдено={found_prices}, не найдено={not_found_prices}")
    
    return data


def save_to_excel(
    data,
    opt_price_file="in/opt_all.xlsx",
    filename="in/products_update.xlsx",
    update_ids=None
):
    logger.info(f"Начинаем сохранение {len(data)} записей в файл {filename}")
    
    if not data:
        logger.error("Нет данных для сохранения")
        return

    # Обогащаем данные ценами из 1С
    data = enrich_products_with_prices(data, opt_price_file)

    # Частичное обновление: добавляем старые записи
    if update_ids and os.path.isfile(filename):
        logger.info("Режим частичного обновления: добавляем старые записи")
        try:
            wb_old = load_workbook(filename, read_only=True)
            ws_old = wb_old.active
            
            if ws_old is not None:
                headers_old = [cell.value for cell in ws_old[1]]
                
                if 'Ozon Product ID' in headers_old:
                    id_idx = headers_old.index('Ozon Product ID')
                    old_records_added = 0
                    
                    # Создаем множество для быстрого поиска обновляемых ID
                    update_set = set(str(uid) for uid in update_ids)
                    
                    # Добавляем записи, которые не обновляются
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if len(row) > id_idx and row[id_idx] is not None:
                            old_id = str(row[id_idx])
                            if old_id not in update_set:
                                # Создаем запись из старых данных
                                old_item = {}
                                for i, header in enumerate(headers_old):
                                    old_item[header] = row[i] if i < len(row) else None
                                data.append(old_item)
                                old_records_added += 1
                    
                    logger.info(f"Добавлено {old_records_added} старых записей, итого: {len(data)}")
            
            wb_old.close()
            
        except Exception as e:
            logger.error(f"Ошибка при добавлении старых записей: {e}")

    # Маппинг полей для правильного отображения в Excel
    field_mapping = {
        "SKU": "SKU",
        "Артикул": "Артикул",
        "Ozon Product ID": "Ozon Product ID",
        "Название товара": "Название товара",
        "Статус товара": "Статус товара",
        "Доступно к продаже по схеме FBS, шт.": "Доступно FBS",
        "Видимость на Ozon": "Видимость",
        "Причины скрытия": "Причины скрытия",
        "Дата создания": "Дата создания",
        "product_link": "Ссылка на товар",
        "base_price": "Базовая цена API",
        "old_price": "Старая цена API",
        "marketing_price": "Маркетинговая цена API",
        "min_price": "Минимальная цена API",
        "Цена": "Цена 1С",
    }
    
    # Обратный маппинг для поиска ключей
    reverse_map = {v: k for k, v in field_mapping.items()}

    # Определяем порядок колонок в итоговом файле
    column_order = [
        "Ozon Product ID", "SKU", "Артикул", "Ссылка на товар", "Название товара",
        "Статус товара", "Видимость", "Причины скрытия", "Базовая цена API",
        "Старая цена API", "Маркетинговая цена API", "Минимальная цена API",
        "Цена 1С", "Доступно FBS", "Дата создания"
    ]

    # Создаём Excel книгу
    logger.debug("Создаем Excel книгу и заполняем данными")
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        logger.error("Не удалось создать лист Excel")
        return
        
    ws.title = "Товары"
    ws.append(column_order)

    # Заполняем данные
    rows_added = 0
    for item in data:
        row = []
        for col in column_order:
            if col == "Ссылка на товар":
                # Генерируем ссылку на товар используя SKU из отчёта
                sku = item.get('SKU', '')
                row.append(f"https://www.ozon.ru/product/{sku}/" if sku and sku != "Н/Д" else "")
                continue
                
            # Получаем значение по ключу
            key = reverse_map.get(col, col)
            val = item.get(key, "")

            # Форматируем ценовые поля
            if col in [
                "Базовая цена API", "Старая цена API",
                "Маркетинговая цена API", "Минимальная цена API",
                "Цена 1С"
            ]:
                if val and val != "Н/Д":
                    try:
                        num = float(str(val).replace(',', '.'))
                        row.append(f"{num:.2f} ₽")
                    except (ValueError, TypeError):
                        row.append(val)
                else:
                    row.append(val)
            elif col == "Доступно FBS":
                # Форматируем количество
                if val:
                    try:
                        row.append(int(val))
                    except (ValueError, TypeError):
                        row.append(val)
                else:
                    row.append(val)
            else:
                row.append(val)
                
        ws.append(row)
        rows_added += 1

    logger.debug(f"Добавлено {rows_added} строк данных в Excel")

    # Автоподбор ширины столбцов
    logger.debug("Настраиваем ширину столбцов")
    for idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(idx)
        
        # Находим максимальную длину текста в колонке
        max_len = 0
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        
        # Устанавливаем ширину с ограничением максимума
        adjusted_width = min((max_len + 2) * 1.2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Создаем директорию если её нет
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # Сохраняем файл
    wb.save(filename)
    logger.info(f"Файл успешно сохранён: {filename}")


def main():
    logger.info("Запуск скрипта получения данных товаров")
    
    parser = argparse.ArgumentParser(description="Скрипт для получения и обработки данных товаров Ozon")
    parser.add_argument("--update-ids", type=str, help="Файл с ID товаров для обновления")
    parser.add_argument("--single-id", type=int, help="ID одного товара для обновления")
    args = parser.parse_args()

    try:
        # Режим обновления конкретных товаров
        if args.single_id or args.update_ids:
            logger.info("Режим частичного обновления товаров")
            
            # Определяем список ID для обновления
            ids = []
            if args.single_id:
                ids = [args.single_id]
                logger.info(f"Обновление одного товара с ID: {args.single_id}")
            else:
                logger.info(f"Загружаем ID товаров из файла: {args.update_ids}")
                try:
                    with open(args.update_ids, "r", encoding='utf-8') as f:
                        ids = [int(line.strip()) for line in f if line.strip().isdigit()]
                    logger.info(f"Загружено {len(ids)} ID товаров для обновления")
                except Exception as e:
                    logger.error(f"Ошибка чтения файла с ID: {e}")
                    return

            if not ids:
                logger.error("Список ID товаров пуст")
                return

            # Создаем и получаем отчет
            code = create_report()
            path = check_report_status(code)
            products = download_report(path)

            # Фильтруем только нужные товары
            logger.info(f"Фильтруем товары по {len(ids)} указанным ID")
            selected_products = []
            ids_set = set(ids)
            
            for product in products:
                product_id_str = product.get("Ozon Product ID")
                if product_id_str and str(product_id_str).strip() and str(product_id_str).strip().isdigit():
                    if int(product_id_str) in ids_set:
                        selected_products.append(product)
            
            logger.info(f"Найдено {len(selected_products)} товаров из {len(ids)} запрошенных")

            if not selected_products:
                logger.warning("Не найдено товаров для обновления")
                return
       
            # Обогащаем данные ценами из API
            logger.info("Обогащаем данные ценами из API")
            processed_products = enrich_products_with_api_data(selected_products)

            # Сохраняем результаты
            output_filename = "in/products_update_single.xlsx"
            logger.info(f"Сохраняем результаты частичного обновления в {output_filename}")
            save_to_excel(
                data=processed_products,
                filename=output_filename,
                update_ids=ids
            )
            
            logger.info("Частичное обновление завершено успешно")
            return

        # Полный режим для всех товаров
        logger.info("Режим полного обновления всех товаров")
        
        # Создаем и получаем отчет
        code = create_report()
        path = check_report_status(code)
        products = download_report(path)

        # Обогащаем данные ценами из API
        logger.info("Обогащаем данные ценами из API")
        processed_products = enrich_products_with_api_data(products)

        logger.info(f"Обработано {len(processed_products)} товаров")

        # Сохраняем результаты полного обновления
        output_filename = "in/products_update_full_vdeeep.xlsx"
        logger.info(f"Сохраняем результаты полного обновления в {output_filename}")
        save_to_excel(
            data=processed_products,
            filename=output_filename,
            update_ids=None
        )
        
        logger.info("Полное обновление завершено успешно")

    except Exception as e:
        logger.exception(f"Критическая ошибка при выполнении скрипта: {e}")
        raise


if __name__ == "__main__":
    main()